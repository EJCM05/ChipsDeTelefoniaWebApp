
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import get_user_model
from django.urls import reverse_lazy
from django.views.generic import (ListView,CreateView,UpdateView,DeleteView,FormView,DetailView,TemplateView)
from core.models import *
from django.shortcuts import render, redirect, get_object_or_404
from .urls import *
from .forms import *
from django.http import JsonResponse
from django.db.models.functions import ExtractMonth
from django.db.models import Count, Q, F, Sum
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import date
from openpyxl.utils import get_column_letter # Importar para obtener la letra de la columna
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from pypdf import PdfReader, PdfWriter
import io
import os
from django.conf import settings
from reportlab.lib.pagesizes import letter # Para obtener las dimensiones de la página si es necesario
from reportlab.lib.colors import white # Para el color blanco
# import qrcode
from PIL import Image
from reportlab.lib.utils import ImageReader
import tempfile
from barcode import Code128 # Importa la clase Code128 directamente
from barcode.writer import ImageWriter
from django.contrib.staticfiles import finders


# Funciones encargadas para el renderizado de paginas

def dashboard_home(request):
    return render(request, "pages/home.html")

def dashboard_clientes(request):
    return render(request, "pages/clientes.html")

def dashboard_ventas(request):
    return render(request, "pages/ventas.html")

def dashboard_estadisticas(request):
    return render(request, "pages/estadisticas.html")

def dashboard_reportes(request):
    return render(request, "pages/reportes.html")

def dashboard_perfil(request):
    return render(request, "pages/perfil.html")

def dashboard_operadoras(request):
    return render(request, "pages/operadoras.html")

# ==========================================================
#               Clases para clientes / buscador
# ==========================================================
class ViewClientes(TemplateView):
    template_name = 'pages/clientes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cedula_buscada = self.request.GET.get('cedula', '').strip()
        
        if cedula_buscada:
            try:
                cliente = Cliente.objects.get(cedula_identidad=cedula_buscada)
                context['cliente'] = cliente
                
                # Obtener todas las SIM Cards asociadas al cliente
                simcards_asociadas = SimCard.objects.filter(id_cliente=cliente)
                context['simcards_asociadas'] = simcards_asociadas
                
                # Obtener todas las ventas asociadas al cliente
                ventas_asociadas = Venta.objects.filter(id_cliente=cliente).order_by('-fecha_venta')
                context['ventas_asociadas'] = ventas_asociadas

            except Cliente.DoesNotExist:
                context['error'] = 'No se encontró ningún cliente con esa cédula.'
            except Exception as e:
                context['error'] = f'Ocurrió un error inesperado: {str(e)}'
        
        return context

# ==========================================================
#               Clases para Lotes
# ==========================================================
class LoteListView(LoginRequiredMixin, ListView):
    model = Lote
    template_name = "pages/home.html"
    context_object_name = "lotes"
    def get_queryset(self):
        # Filtra los lotes para mostrar solo los que pertenecen al usuario actual.
        # `self.request.user` es el usuario que ha iniciado sesión.
        return Lote.objects.filter(id_usuario_propietario=self.request.user)

# Vista para crear un nuevo lote
class LoteCreateView(LoginRequiredMixin, CreateView):
    model = Lote
    template_name = "components/formularios/lotes/lote_form.html"
    form_class = LoteForm
    success_url = reverse_lazy("dashboard:dashboard_home")

    # Sobreescribe el método form_valid para asignar el usuario actual al lote
    def form_valid(self, form):
        Usuario = get_user_model()
        # Asigna el usuario actual como propietario del lote
        form.instance.id_usuario_propietario = self.request.user
        return super().form_valid(form)


# Vista para actualizar un lote existente
class LoteUpdateView(LoginRequiredMixin, UpdateView):
    model = Lote
    template_name = "components/formularios/lotes/lote_form.html"
    form_class = LoteForm
    success_url = reverse_lazy("dashboard:dashboard_home")


# Vista para eliminar un lote
class LoteDeleteView(LoginRequiredMixin, DeleteView):
    model = Lote
    template_name = "components/confirmaciones/lotes/lote_confirm_delete.html"
    success_url = reverse_lazy("dashboard:dashboard_home")
    
    
# ============================================================
#            Clases para SIMCards
# ============================================================

# Vista para listar las SIMCards de un lote específico
class SimCardListView(LoginRequiredMixin, ListView):
    model = SimCard
    template_name = "pages/simcards_list.html" # Crea este template
    context_object_name = "simcards"

    def get_queryset(self):
        # Filtra las SIMCards por el lote pasado en la URL
        lote_pk = self.kwargs['pk']
        return SimCard.objects.filter(id_lote__pk=lote_pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agrega el lote al contexto para usarlo en el template
        context['lote'] = Lote.objects.get(pk=self.kwargs['pk'])
        return context

# Vista para crear una nueva SIMCard en un lote específico
class SimCardCreateView(LoginRequiredMixin, CreateView):
    model = SimCard
    template_name = "components/formularios/simscards/simscards_form.html"
    form_class = SimCardForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['lote'] = Lote.objects.get(pk=self.kwargs['pk'])
        return context

    def get_success_url(self):
        lote_pk = self.kwargs['pk']
        return reverse_lazy('dashboard:simcard_list', kwargs={'pk': lote_pk})

    def form_valid(self, form):
        lote_pk = self.kwargs['pk']
        lote = Lote.objects.get(pk=lote_pk)
        
        # Asigna el lote y el usuario propietario
        form.instance.id_lote = lote
        # form.instance.id_usuario_propietario = self.request.user # Si existe este campo
        
        return super().form_valid(form)

# Vista para actualizar una SIMCard existente
class SimCardUpdateView(LoginRequiredMixin, UpdateView):
    model = SimCard
    template_name = "components/formularios/simscards/simscards_form.html"
    form_class = SimCardForm
    
    def get_success_url(self):
        # Redirige a la lista de SIMCards del lote actual
        simcard = self.get_object()
        lote_pk = simcard.id_lote.pk
        return reverse_lazy('dashboard:simcard_list', kwargs={'pk': lote_pk})

# Vista para eliminar una SIMCard
class SimCardDeleteView(LoginRequiredMixin, DeleteView):
    model = SimCard
    template_name = "components/confirmaciones/simscards/simcard_confirm_delete.html"
    
    def get_success_url(self):
        # Redirige a la lista de SIMCards del lote actual
        simcard = self.get_object()
        lote_pk = simcard.id_lote.pk
        return reverse_lazy('dashboard:simcard_list', kwargs={'pk': lote_pk})
    
# ======================================================================
#               Clases para Operadoras 
# ======================================================================
    
class OperadoraListView(LoginRequiredMixin, ListView):
    model = Operadora
    template_name = "components/tablas/operadoras/operadora_list.html"
    context_object_name = "operadoras"

# Vista para crear una nueva operadora
class OperadoraCreateView(LoginRequiredMixin, CreateView):
    model = Operadora
    template_name = "components/formularios/operadoras/operadoras_form.html"
    form_class = OperadoraForm
    success_url = reverse_lazy("dashboard:dashboard_operadoras")

# Vista para actualizar una operadora existente
class OperadoraUpdateView(LoginRequiredMixin, UpdateView):
    model = Operadora
    template_name = "components/formularios/operadoras/operadoras_form.html"
    form_class = OperadoraForm
    success_url = reverse_lazy("dashboard:dashboard_operadoras")

# Vista para eliminar una operadora
class OperadoraDeleteView(LoginRequiredMixin, DeleteView):
    model = Operadora
    template_name = "components/confirmaciones/operadoras/operadoras_confirm_delete.html"
    success_url = reverse_lazy("dashboard:dashboard_operadoras")
    
# ==========================================================================
                    # Clases ventas
# ==========================================================================


class VentaDeleteView(DeleteView):
    model = Venta
    template_name = 'components/confirmaciones/ventas/venta_confirm_delete.html'
    success_url = reverse_lazy('dashboard:dashboard_ventas')
    
class GenerarVentaView(LoginRequiredMixin, CreateView):
    model = Venta
    form_class = VentaForm
    template_name = 'components/formularios/ventas/formulario_ventas.html'
    success_url = reverse_lazy('dashboard:dashboard_ventas')

    def form_valid(self, form):
        self.object = form.save(user=self.request.user) 
        return redirect(self.get_success_url())

class EditarVentaView(LoginRequiredMixin, UpdateView):
    model = Venta
    form_class = VentaForm
    template_name = 'components/formularios/ventas/formulario_ventas.html'
    success_url = reverse_lazy('dashboard:dashboard_ventas')

    def form_valid(self, form):
        form.save()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

# Vista para listar las ventas (reportes)
class ListaVentasView(LoginRequiredMixin, ListView):
    model = Venta
    template_name = 'pages/ventas.html'
    context_object_name = 'ventas'

    def get_queryset(self):
        # Opcional: Muestra solo las ventas del usuario actual si es supervisor
        if self.request.user.rol == 'supervisor':
            return Venta.objects.filter(id_usuario_propietario=self.request.user).order_by('-fecha_venta')
        return Venta.objects.all().order_by('-fecha_venta')

# Vista de detalle para ver una factura/reporte individual
class DetalleVentaView(LoginRequiredMixin, DetailView):
    model = Venta
    template_name = 'components/tablas/ventas/detalles_reportes.html'
    context_object_name = 'venta'
    
    def get_queryset(self):
        # Asegura que un supervisor solo pueda ver sus propias ventas
        if self.request.user.rol == 'supervisor':
            return Venta.objects.filter(id_usuario_propietario=self.request.user)
        return Venta.objects.all()
    
# ==========================================================================
                # Clase Reportes de ventas
# ==========================================================================

class ListaReportesView(LoginRequiredMixin, ListView):
    model = Venta
    template_name = 'pages/reportes.html'
    context_object_name = 'ventas'
    paginate_by = 10 # Número de ventas por página

    def get_queryset(self):
        # Asegura que las ventas se muestren de la más reciente a la más antigua
        # y filtra por el usuario si es un supervisor
        queryset = Venta.objects.all().order_by('-fecha_venta')

        # Restringe los reportes si el usuario no es un administrador
        if self.request.user.rol == 'supervisor':
            queryset = queryset.filter(id_usuario_propietario=self.request.user)

        return queryset


# # =====================================================
            # Funciones
# # =====================================================


def generar_contrato_pdf(request, pk):
    try:
        # 1. Obtener la venta
        venta = get_object_or_404(Venta, pk=pk)

        # 2. Configurar la plantilla y un buffer para el contenido a añadir
        template_path = finders.find('pdf/base.pdf')

        if not template_path:
            # Si finders.find() no encuentra el archivo, devuelve un 500 con un mensaje útil.
            # En producción, esto es útil para los logs de error.
            return HttpResponse("Error 500: El archivo de plantilla PDF no se encuentra en el servidor.", status=500)
        
        buffer_output = io.BytesIO()
        p = canvas.Canvas(buffer_output)
        
        # Función auxiliar para dibujar un rectángulo blanco y luego el texto encima
        def draw_text_with_white_background(x, y, text, font_name="Helvetica", font_size=10, padding_x=2, padding_y=2):
            # Establece la fuente para calcular el ancho del texto
            p.setFont(font_name, font_size)
            
            # Calcular el ancho del texto
            text_width = p.stringWidth(text, font_name, font_size)
            
            # Calcular la altura del texto (más preciso)
            # Una estimación aproximada para la altura de la línea es 1.2 veces el tamaño de la fuente
            text_height = font_size * 1.2
            
            # Dibuja un rectángulo blanco sin borde
            p.setFillColor(white)
            p.setStrokeColor(white)
            p.rect(x - padding_x, y - padding_y, text_width + (2 * padding_x), text_height + (2 * padding_y), stroke=0, fill=1)
            
            # Restaura el color del texto a negro y dibuja el texto
            p.setFillColorRGB(0,0,0)
            p.drawString(x, y, text)

        # Coordenadas y tamaños de los campos de texto
        draw_text_with_white_background(445, 720, f"{venta.id_simcard.numero_telefono}", font_size=15)
        draw_text_with_white_background(30, 632, f"{venta.id_cliente.primer_nombre} {venta.id_cliente.segundo_nombre} {venta.id_cliente.primer_apellido} {venta.id_cliente.segundo_apellido}", font_size=8)
        draw_text_with_white_background(512, 632, f"V{venta.id_cliente.cedula_identidad}", font_size=10)
        draw_text_with_white_background(40, 198, f"V{venta.id_cliente.cedula_identidad}", font_size=10)
        draw_text_with_white_background(165, 198, f"{venta.fecha_venta.strftime('%d/%m/%Y')}", font_size=10)
        draw_text_with_white_background(37, 603, f"{venta.id_cliente.telefono}", font_size=8)
        draw_text_with_white_background(37, 575, f"{venta.id_cliente.direccion}", font_size=8)
        
        # Coordenadas y tamaño del código de barras
        barcode_x_coord = 240    # Coordenada X para el código de barras
        barcode_y_coord = 770     # Coordenada Y para el código de barras (en la parte inferior)

        # Generar el código de barras con el código de la SIM
        # Usamos Code128 porque es un formato común y versátil
        code128 = Code128(venta.id_simcard.codigo, writer=ImageWriter()) # Se utiliza la clase directamente
        
        # Usa un buffer de memoria para la imagen del código de barras
        buffer_barcode = io.BytesIO()
        code128.write(buffer_barcode)
        buffer_barcode.seek(0)
        
        # Abre el buffer como una imagen de Pillow
        img_barcode = Image.open(buffer_barcode)
        
        # Crea un objeto ImageReader de reportlab a partir de la imagen de Pillow
        img_reader = ImageReader(img_barcode)
        
        # Dibuja la imagen del código de barras en el PDF
        p.drawImage(img_reader, barcode_x_coord, barcode_y_coord, width=100, height=50)

        p.showPage()
        p.save()
        
        # 4. Combinar el PDF de la plantilla con el PDF de los datos
        buffer_output.seek(0)
        
        existing_pdf = PdfReader(template_path)
        new_pdf = PdfReader(buffer_output)
        
        output = PdfWriter()
        
        if len(existing_pdf.pages) > 0:
            page = existing_pdf.pages[0]
            page.merge_page(new_pdf.pages[0])
            output.add_page(page)
        else:
            output.add_page(new_pdf.pages[0])
        
        # 5. Devolver el PDF combinado
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="contrato_venta_{venta.id_simcard.numero_telefono}.pdf"'
        
        output.write(response)

        return response
    except Exception as e:
        # Si ocurre cualquier otro error, este bloque lo capturará
        # y te dirá exactamente qué está pasando en el log de Render
        return HttpResponse(f"Error 500: Un error inesperado ha ocurrido. Detalles: {e}", status=500)


# Funcion para generar excel con la muestra de ventas
def generar_reporte_lotes_excel(request):
    # Obtener el año y el mes de los parámetros GET
    year = request.GET.get('year')
    month = request.GET.get('month')

    if not year or not month:
        # Si no se ha seleccionado un mes, simplemente mostramos el formulario
        return render(request, 'estadisticas/reporte_lotes.html')

    try:
        year = int(year)
        month = int(month)
    except (ValueError, TypeError):
        # Manejamos errores si los valores no son números
        return render(request, 'estadisticas/reporte_lotes.html', {'error': 'Año o mes no válido.'})

    # Crear un nuevo libro de trabajo de Excel
    workbook = openpyxl.Workbook()
    # Eliminar la hoja por defecto que se crea
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Filtrar los lotes
    lotes = Lote.objects.all()

    # Recorrer cada lote para crear una hoja
    for lote in lotes:
        # Crear una nueva hoja con el nombre del lote
        sheet = workbook.create_sheet(title=lote.nombre_lote)

        # Encabezados de la hoja
        sheet.merge_cells('A1:D1')
        title_cell = sheet['A1']
        title_cell.value = f"Reporte del Lote '{lote.nombre_lote}' - Mes {month}/{year}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        # Encabezados de las columnas para SIM Cards
        headers_simcard = [
            "Código de SIM Card",
            "Número de Teléfono",
            "Estado",
            "Fecha de Creación"
        ]
        
        # Escribir los encabezados de SIM Cards
        for col_num, header in enumerate(headers_simcard, 1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Filtrar SIM Cards por lote y mes/año
        simcards = SimCard.objects.filter(
            id_lote=lote,
            created_at__year=year,
            created_at__month=month
        ).order_by('created_at')

        row_num = 4
        if simcards.exists():
            for simcard in simcards:
                sheet.cell(row=row_num, column=1, value=simcard.codigo)
                sheet.cell(row=row_num, column=2, value=simcard.numero_telefono if simcard.numero_telefono else "No asignado")
                sheet.cell(row=row_num, column=3, value=simcard.get_estado_display())
                sheet.cell(row=row_num, column=4, value=simcard.created_at.strftime("%Y-%m-%d"))
                row_num += 1
        else:
            sheet.cell(row=row_num, column=1, value="No se encontraron SIM Cards en este mes.")
            sheet.cell(row=row_num, column=1).font = Font(italic=True)
            sheet.merge_cells(f'A{row_num}:D{row_num}')
            row_num += 1

        # Separador para las Ventas
        row_num += 2
        
        # Encabezados de las columnas para Ventas
        sheet.merge_cells(f'A{row_num}:D{row_num}')
        ventas_title_cell = sheet[f'A{row_num}']
        ventas_title_cell.value = "Ventas Realizadas"
        ventas_title_cell.font = Font(bold=True, size=12)
        ventas_title_cell.alignment = Alignment(horizontal='center')
        
        row_num += 1
        headers_venta = [
            "Fecha de Venta",
            "Monto Total",
            "Cédula del Cliente",
            "Código de SIM Card"
        ]
        
        for col_num, header in enumerate(headers_venta, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            
        row_num += 1

        # Filtrar Ventas por lote y mes/año
        ventas = Venta.objects.filter(
            id_simcard__id_lote=lote,
            fecha_venta__year=year,
            fecha_venta__month=month
        ).order_by('fecha_venta')

        if ventas.exists():
            for venta in ventas:
                sheet.cell(row=row_num, column=1, value=venta.fecha_venta.strftime("%Y-%m-%d %H:%M"))
                sheet.cell(row=row_num, column=2, value=float(venta.monto_total))
                sheet.cell(row=row_num, column=3, value=venta.id_cliente.cedula_identidad if venta.id_cliente else "Cliente eliminado")
                sheet.cell(row=row_num, column=4, value=venta.id_simcard.codigo if venta.id_simcard else "SIM Card eliminada")
                row_num += 1
        else:
            sheet.cell(row=row_num, column=1, value="No se encontraron ventas en este mes.")
            sheet.cell(row=row_num, column=1).font = Font(italic=True)
            sheet.merge_cells(f'A{row_num}:D{row_num}')

        # SOLUCIÓN: Ajustar el ancho de las columnas de forma segura
        for col in range(1, 5): # Iterar sobre las columnas A, B, C, D
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, sheet.max_row + 1):
                cell = sheet[f"{column_letter}{row}"]
                # Evitar las celdas fusionadas
                if cell.coordinate not in sheet.merged_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Reporte_Lotes_{year}_{month}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Guardar el libro de trabajo en la respuesta HTTP
    workbook.save(response)
    return response


# Funcion auxiliar para el envio de datos a la grafica 
def obtener_ventas(request, year):
    # Filtra las ventas por el año solicitado
    ventas = Venta.objects.filter(fecha_venta__year=year)

    # Agrega la cantidad de chips y el monto total por mes
    datos_agregados = ventas.annotate(
        mes=ExtractMonth('fecha_venta')
    ).values('mes').annotate(
        chips_vendidos=Count('id_simcard'),
        monto_total=Sum('monto_total')
    ).order_by('mes')

    # Prepara los datos para la respuesta JSON
    labels = [
        'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
        'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic'
    ]
    
    chips_vendidos = [0] * 12
    monto_total = [0] * 12

    for dato in datos_agregados:
        # Los meses de Django son 1-12, los índices de la lista son 0-11
        mes_index = dato['mes'] - 1
        chips_vendidos[mes_index] = dato['chips_vendidos']
        monto_total[mes_index] = float(dato['monto_total'])

    response_data = {
        'labels': labels,
        'chipsVendidos': chips_vendidos,
        'montoTotal': monto_total
    }
    
    return JsonResponse(response_data)


def get_simcards_by_lote(request):
    """
    Vista que devuelve una lista de SIM Cards disponibles
    (estado 'inactiva') para un lote específico, en formato JSON.
    """
    if request.method == 'GET' and 'lote_id' in request.GET:
        try:
            lote_id = request.GET.get('lote_id')
            # Filtra por lote, excluyendo las que ya están vendidas,
            # pero permitiendo la de la venta actual si está en la URL.
            simcards = SimCard.objects.filter(id_lote=lote_id).exclude(estado='vendida')

            # Si se está editando una venta, incluye la simcard de la venta actual
            venta_id = request.GET.get('venta_id')
            if venta_id:
                try:
                    venta = Venta.objects.get(pk=venta_id)
                    simcards = simcards | SimCard.objects.filter(pk=venta.id_simcard.pk)
                except Venta.DoesNotExist:
                    pass

            data = [
                {
                    'id': simcard.id,
                    'codigo': simcard.codigo,
                    'estado': simcard.estado
                }
                for simcard in simcards
            ]
            print(data)
            return JsonResponse({'simcards': data})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)